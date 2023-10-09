import openpyxl

def get_token_column(sheet, token, is_column = True):
    if is_column:
        for column in sheet.iter_cols():
            for cell in column:
                if token == str(cell.value):
                    return cell.column
    else:
        for row in sheet.iter_rows():
            for cell in row:
                if token == str(cell.value):
                    return cell.row
    return -1

wb = openpyxl.load_workbook('e:\Code\input.xlsx')

sheet = wb.active

token = '现货库存'

#获取地区信息
area = []
rdc_rows = []
for cell in sheet[1]:
    if token in cell.value and not('全国' in cell.value):
        area.append(cell.value.replace(token,''))


# 获取第一行的所有单元格的值
row_values = []
row_values_index =[]
for cell in sheet[1]:
    if area[0] in cell.value:
        row_values.append(cell.value.replace(area[0], ''))

print(row_values)


#get rdc in use



#get in use sku rows
sku_col = get_token_column(sheet,'SKU',is_column = True)
sku_price_col = get_token_column(sheet, '全国采购价',is_column = True)
sku_name_col = get_token_column(sheet, '商品名称',is_column = True)
n = get_token_column(sheet, '全国现货库存', True)

skus_index = []
skus_name =[]
skus = []
skus_price = []


list_sheet = list(sheet.columns)
print(len(list_sheet))
for row in range(1,len(list_sheet[n-1])):
    if int(list_sheet[n-1][row].value) > 0:
        skus_index.append(row)
        skus.append(list_sheet[sku_col-1][row].value)
        skus_price.append(list_sheet[sku_price_col-1][row].value)
        skus_name.append(list_sheet[sku_name_col-1][row].value)

print(skus) 
new_sheet=[]

first_row = []
first_row.append('SKU')
first_row.append('商品名称')
first_row.append('供货价')
first_row.append('地区')


first_row.extend(row_values)

new_sheet.append(first_row)

for i in range(0,len(skus)):
    for j in range(0,len(area)):
        row =[]
        row.append(skus[i])
        row.append(skus_name[i])
        row.append(skus_price[i])
        row.append(area[j])

        iter=0
        for col in range(0,len(row_values)):
            token = area[j] + row_values[col]
            if col == 0:
                iter = get_token_column(sheet, token)
            else:
                iter +=1
            row.append(list_sheet[iter-1][skus_index[i]].value)
        new_sheet.append(row)
        #print(row)
        
print(new_sheet[0])
print(new_sheet[1])

def save_list_to_xlsx(file_name, sheet_name, data_list):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()

    # 创建一个新的工作表
    sheet = wb.active
    sheet.title = sheet_name

    # 将列表中的数据逐个写入工作表的单元格中
    for row_index, row_data in enumerate(data_list, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index).value = cell_value

    # 保存工作簿为xlsx文件
    wb.save(file_name)


# 调用函数保存为xlsx文件
save_list_to_xlsx('e:\Code\output.xlsx', 'sheet1', new_sheet)